#!/usr/bin/env python3
"""
Script xuất danh sách câu hỏi ra file Excel (.xlsx) với format đẹp
"""

import re
import sys
import os
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Dang cai dat thu vien openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

def parse_markdown_tracking(file_path):
    """Parse file markdown tracking để extract câu hỏi từ bảng"""
    questions = []
    in_table = False
    header_processed = False
    header_cols = 0  # Số cột trong header
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    lines = content.split('\n')
    
    for i, line in enumerate(lines):
        # Bắt đầu bảng - lưu số cột của header
        if '|' in line and 'Nhóm nội dung' in line:
            in_table = True
            header_processed = True
            header_cols = len([c for c in line.split('|') if c.strip()])
            continue
        
        # Kết thúc bảng
        if in_table and (line.strip() == '' or (not '|' in line and line.strip() != '')):
            if line.strip() != '':
                in_table = False
            continue
        
        # Xử lý dòng trong bảng (bỏ qua header và separator)
        if in_table and header_processed and '|' in line:
            # Bỏ qua dòng separator (chứa ---)
            if '---' in line:
                continue
            
            # Parse các cột
            cols = [col.strip() for col in line.split('|')]
            # Bỏ phần tử đầu và cuối rỗng (do format markdown table)
            if len(cols) > 2:
                cols = cols[1:-1] if cols[0] == '' and cols[-1] == '' else cols
            
            # Xác định số cột dữ liệu thực tế (bỏ qua phần tử rỗng)
            data_cols = [c for c in cols if c]
            
            # Header có 5 cột: Nhóm, Câu hỏi, Câu trả lời tham khảo, Ngày, Làm rõ
            # Dòng dữ liệu có thể có 4 cột (không có Ngày) hoặc 5 cột (có Ngày)
            if len(data_cols) >= 4:
                nhom = data_cols[0] if len(data_cols) > 0 else ''
                cau_hoi = data_cols[1] if len(data_cols) > 1 else ''
                tra_loi_tham_khao = data_cols[2] if len(data_cols) > 2 else ''
                
                # Kiểm tra xem có cột "Ngày" không
                # Nếu có 5 cột dữ liệu thì cột thứ 4 là "Ngày", cột thứ 5 là "Làm rõ"
                # Nếu có 4 cột dữ liệu thì cột thứ 4 là "Làm rõ" (không có "Ngày")
                if len(data_cols) >= 5:  # Có cột Ngày
                    ngay = data_cols[3] if len(data_cols) > 3 else ''
                    lam_ro = data_cols[4] if len(data_cols) > 4 else ''
                else:  # Không có cột Ngày (format cũ)
                    ngay = ''
                    lam_ro = data_cols[3] if len(data_cols) > 3 else ''
                
                # Xác định trạng thái từ checkbox
                is_answered = '✅' in lam_ro or '[x]' in lam_ro.lower() or '[X]' in lam_ro.lower()
                
                if cau_hoi:  # Chỉ thêm nếu có câu hỏi
                    questions.append({
                        'nhom': nhom,
                        'cau_hoi': cau_hoi,
                        'tra_loi_tham_khao': tra_loi_tham_khao,
                        'ngay': ngay,
                        'lam_ro': 'Đã làm rõ' if is_answered else 'Chưa làm rõ'
                    })
    
    # Nếu không tìm thấy bảng, fallback về cách cũ (parse checkbox)
    if not questions:
        current_group = None
        question_num = 1
        
        for line in lines:
            # Nhóm câu hỏi (##)
            if line.startswith('## ') and not line.startswith('###'):
                current_group = line.replace('##', '').strip()
                continue
            
            # Câu hỏi với checkbox
            if line.strip().startswith('- [ ]') or line.strip().startswith('- [x]') or line.strip().startswith('- [X]'):
                question_text = line.strip()
                is_answered = '[x]' in question_text.lower() or '[X]' in question_text.lower()
                question_text = re.sub(r'^- \[[xX ]\]\s*', '', question_text)
                
                if question_text and current_group:
                    questions.append({
                        'nhom': current_group,
                        'cau_hoi': question_text,
                        'tra_loi_tham_khao': '',
                        'ngay': '',
                        'lam_ro': 'Đã làm rõ' if is_answered else 'Chưa làm rõ'
                    })
                    question_num += 1
    
    return questions

def calculate_column_width(text, min_width=10, max_width=80):
    """Tính độ rộng cột dựa trên nội dung"""
    if not text:
        return min_width
    
    # Đếm số ký tự (tiếng Việt tính 1.2 ký tự)
    lines = str(text).split('\n')
    max_line_length = max(len(line) for line in lines) if lines else 0
    
    # Tính độ rộng (khoảng 1.1 ký tự cho mỗi đơn vị width trong Excel)
    width = max(min_width, min(max_line_length * 1.1 + 2, max_width))
    return width

def export_to_excel(questions, output_file):
    """Export questions ra Excel với format đẹp"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách câu hỏi"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Data alignment (wrap text cho tất cả cells)
    data_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    
    # Header
    headers = ['Nhóm nội dung', 'Câu hỏi', 'Câu trả lời tham khảo', 'Ngày', 'Làm rõ']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    max_widths = [0] * len(headers)  # Lưu độ rộng tối đa của mỗi cột
    
    for row_idx, q in enumerate(questions, start=2):
        values = [
            q.get('nhom', ''),
            q.get('cau_hoi', ''),
            q.get('tra_loi_tham_khao', ''),
            q.get('ngay', ''),
            q.get('lam_ro', 'Chưa làm rõ')
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Cập nhật độ rộng tối đa
            width = calculate_column_width(value)
            if width > max_widths[col_idx - 1]:
                max_widths[col_idx - 1] = width
    
    # Set column widths
    # Nhóm nội dung: 25
    # Câu hỏi: 60 (câu hỏi dài)
    # Câu trả lời tham khảo: 70 (nội dung tham khảo dài)
    # Ngày: 12
    # Làm rõ: 15
    column_widths = [25, 60, 70, 12, 15]
    
    for col_idx, width in enumerate(column_widths, 1):
        # Sử dụng độ rộng tính toán hoặc giá trị mặc định, lấy giá trị lớn hơn
        calculated_width = max_widths[col_idx - 1]
        final_width = max(width, calculated_width) if calculated_width > 0 else width
        ws.column_dimensions[get_column_letter(col_idx)].width = final_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(questions) + 1}"
    
    # Save file
    try:
        wb.save(output_file)
        try:
            print(f"SUCCESS: Da xuat {len(questions)} cau hoi ra file: {output_file}")
            print(f"   - Wrap text: Da bat cho tat ca cac cell")
            print(f"   - Column width: Da tu dong dieu chinh")
        except:
            print(f"SUCCESS: Da xuat {len(questions)} cau hoi ra file Excel")
    except PermissionError:
        print(f"ERROR: Khong the ghi file '{output_file}'")
        print(f"   File dang duoc mo trong Excel hoac ung dung khac.")
        print(f"   Vui long dong file va chay lai script.")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Loi khi luu file: {e}")
        sys.exit(1)

def main():
    if len(sys.argv) < 2:
        print("Usage: python export_to_excel.py <tracking_file.md> [output.xlsx]")
        sys.exit(1)
    
    tracking_file = Path(sys.argv[1])
    if not tracking_file.exists():
        print(f"ERROR: File khong ton tai: {tracking_file}")
        sys.exit(1)
    
    # Output file
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
        # Đảm bảo extension là .xlsx
        if not output_file.endswith('.xlsx'):
            output_file = str(Path(output_file).with_suffix('.xlsx'))
    else:
        # Tự động tạo tên file
        date_str = datetime.now().strftime('%Y%m%d')
        output_file = f"questions_{date_str}.xlsx"
    
    # Parse và export
    questions = parse_markdown_tracking(tracking_file)
    export_to_excel(questions, output_file)

if __name__ == '__main__':
    main()

